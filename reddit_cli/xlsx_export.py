"""XLSX export utilities for Reddit CLI."""

from io import BytesIO

from reddit_cli.reddit.models import Comment, Post, Subreddit


def _check_openpyxl() -> None:
    """Check if openpyxl is installed.

    Raises:
        ImportError: If openpyxl is not installed
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX export. "
            "Please install it with: pip install openpyxl"
        )


def posts_to_xlsx(posts: list[Post], sheet_name: str = "Posts") -> bytes:
    """Convert a list of posts to XLSX format.

    Args:
        posts: List of Post objects
        sheet_name: Name of the worksheet

    Returns:
        XLSX file content as bytes

    Raises:
        ImportError: If openpyxl is not installed
    """
    _check_openpyxl()

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers
    headers = ["id", "title", "author", "subreddit", "score", "num_comments", "url", "permalink", "created_utc", "selftext"]
    ws.append(headers)

    # Data rows
    for post in posts:
        ws.append([
            post.id,
            post.title,
            post.author,
            post.subreddit,
            post.score,
            post.num_comments,
            post.url,
            post.permalink,
            post.created_utc,
            post.selftext,
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def comments_to_xlsx(comments: list[Comment], sheet_name: str = "Comments") -> bytes:
    """Convert a list of comments to XLSX format.

    Args:
        comments: List of Comment objects
        sheet_name: Name of the worksheet

    Returns:
        XLSX file content as bytes

    Raises:
        ImportError: If openpyxl is not installed
    """
    _check_openpyxl()

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers
    headers = ["id", "author", "body", "score", "created_utc", "parent_id", "link_id", "depth"]
    ws.append(headers)

    # Flatten comments and add rows
    def flatten_and_add(comments_list: list[Comment]) -> None:
        for comment in comments_list:
            ws.append([
                comment.id,
                comment.author,
                comment.body,
                comment.score,
                comment.created_utc,
                comment.parent_id,
                comment.link_id,
                comment.depth,
            ])
            if comment.replies:
                flatten_and_add(comment.replies)

    flatten_and_add(comments)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def subreddits_to_xlsx(subreddits: list[Subreddit], sheet_name: str = "Subreddits") -> bytes:
    """Convert a list of subreddits to XLSX format.

    Args:
        subreddits: List of Subreddit objects
        sheet_name: Name of the worksheet

    Returns:
        XLSX file content as bytes

    Raises:
        ImportError: If openpyxl is not installed
    """
    _check_openpyxl()

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers
    headers = ["id", "display_name", "title", "description", "subscribers", "active_users"]
    ws.append(headers)

    # Data rows
    for subreddit in subreddits:
        ws.append([
            subreddit.id,
            subreddit.display_name,
            subreddit.title,
            subreddit.description,
            subreddit.subscribers,
            subreddit.active_users,
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
